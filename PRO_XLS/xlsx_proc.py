# 针对 03版excel（xls结尾的），我们可以使用xlrd读，xlwt包来写 
# 针对 07版excel（xlsx结尾的）,我们可以使用openpyxl来操作读写excel 


#读取excel使用(支持03)
import xlrd
#写入excel使用(支持03)
import xlwt
#读取execel使用(支持07)
from openpyxl import Workbook
#写入excel使用(支持07)
from openpyxl import load_workbook

# import json

# a='{"1":["张三",150,120,100],"2":["李四",90,99,95],"3":["王五",60,66,68]}'
# 
# b=json.loads(a, encoding='utf-8')
# print(b)
def showexcel(path):
    workbook=xlrd.open_workbook(path)
    sheets=workbook.sheet_names();
    #多个sheet时，采用下面的写法打印
    #for sname in sheets:
        #print(sname)
    worksheet=workbook.sheet_by_name(sheets[0])
    #print(worksheet)
    #nrows=worksheet.nrows
    #nclows=worksheet.ncols
    for i in range(0,worksheet.nrows):

        for j in range(0,worksheet.ncols):
            print(worksheet.cell_value(i,j),"\t",end="")

        print()

def getrowsexcel(path):
    workbook=xlrd.open_workbook(path)
    sheets=workbook.sheet_names();
    #多个sheet时，采用下面的写法打印
    #for sname in sheets:
        #print(sname)
    worksheet=workbook.sheet_by_name(sheets[0])
    #print(worksheet)
    rows=list(worksheet.get_rows())
    rows=[[rows[j][i].value for i in range(0,len(rows[j])-1) ] for j in range(0,len(rows)-1)]
    return rows
    #nclows=worksheet.ncols
def writerowsexcel07(path,row):

    wb=Workbook(path)
    #sheet=wb.add_sheet("xlwt数据测试表")
    sheet=wb.create_sheet("sheet1",0)
    for i in row:
        sheet.append(i)
    wb.save(path)
def writeexcel03(path):

    wb=xlwt.Workbook()
    sheet=wb.add_sheet("xlwt数据测试表")
    value = [["名称", "hadoop编程实战", "hbase编程实战", "lucene编程实战"], ["价格", "52.3", "45", "36"], ["出版社", "机械工业出版社", "人民邮电出版社", "华夏人民出版社"], ["中文版式", "中", "英", "英"]]
    for i in range(0,4):
        for j in range(0,len(value[i])):
            sheet.write(i,j,value[i][j])
    wb.save(path)
    print("写入数据成功！")

def writeexcel07(path):

    wb=Workbook()
    #sheet=wb.add_sheet("xlwt数据测试表")
    sheet=wb.create_sheet("xlwt数据测试表",0)

    value = [["名称", "hadoop编程实战", "hbase编程实战", "lucene编程实战"], ["价格", "52.3", "45", "36"], ["出版社", "机械工业出版社", "人民邮电出版社", "华夏人民出版社"], ["中文版式", "中", "英", "英"]]
    for i in range(0,4):
            sheet.append(value[i])
    sheet.cell(row = 1,column= 2).value="温度"
    wb.save(path)
    print("写入数据成功！")


def read07excel(path):
    wb2=load_workbook(path)
    #print(wb2.get_sheet_names())
    ws=wb2.get_sheet_by_name("详单一")
    row=ws.get_highest_row()
    col=ws.get_highest_column()
    print("列数: ",ws.get_highest_column())
    print("行数: ",ws.get_highest_row())

    for i  in range(0,row):
        for j in range(0,col):
            print(ws.rows[i][j].value,"\t\t",end="")

        print()

def find_row(rows,findname):
    for row in rows:
        for i in row:
            try:
                if i.find(findname)!=-1:
                    #print(row)
                    return row
                    break
            except:
                pass
    return [findname,'##未找到###']
if __name__ == '__main__':
    read03path=r"国浩保险法律部通讯录.xls"
    writepath=r'找到的通讯录的名单.xlsx'
    with open(u'名单列表.txt') as fp:
        name_list=fp.readlines()
    print(name_list)
    '''读取excel'''
    rows=getrowsexcel(read03path)
    finded_row=list(map(lambda x: find_row(rows,x.strip()),name_list))
    finded_row.insert(0, rows[0])
#     filter_row=[i for i in finded_row if i!=None]
#     print(filter_row)
    if finded_row == None:
        print('失败 未找到文本')
    else:
        writerowsexcel07(writepath,finded_row)
    print('done')
